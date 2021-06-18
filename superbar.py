# -*- coding: utf-8 -*-
"""
Created on Mon May 31 17:27:44 2021

@author: TC
"""

# import numpy as np
from pprint import pprint
from openpyxl import *
import datetime
from copy import copy
# from openpyxl_modules import *
from openpyxl.styles.borders import Border, Side
import openpyxl
from openpyxl import drawing 
import shutil
import os
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

def get_combis(besoins_shaped : list, besoins_max : list) -> list :
    combis = list()
    for b in range(len(besoins_shaped)) :
        if besoins_shaped[b][0][0] == "Joint" :
            # combis.append([])
            combi = list()
            combi = [k for k in range(len(besoins_shaped[b]))]
            # combi[-1].append([k for k in range(len(besoins_shaped[b]))])
            # print(combi[-1])
            # pprint(besoins_shaped[b])
            combis.append([combi])
        else :
            combis.append([])
            combi = list()
            n = len(besoins_shaped[b])
            k = besoins_max[b][1]
            
            if n >= 1 : combi.append([])
            for i1 in range(n) :
                combi[-1].append([i1])
            if n >= 2 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    combi[-1].append([i1, i2])
            if n >= 3 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        combi[-1].append([i1, i2, i3])
            if n >= 4 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            combi[-1].append([i1, i2, i3, i4])
            if n >= 5 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                combi[-1].append([i1, i2, i3, i4, i5])
            if n >= 6 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                for i6 in range(i5 + 1, n) :
                                    combi[-1].append([i1, i2, i3, i4, i5, i6])
            if n >= 7 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                for i6 in range(i5 + 1, n) :
                                    for i7 in range(i6 + 1, n) :
                                        combi[-1].append([i1, i2, i3, i4, i5, i6, i7])
            if n >= 8 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                for i6 in range(i5 + 1, n) :
                                    for i7 in range(i6 + 1, n) :
                                        for i8 in range(i7 + 1, n) :
                                            combi[-1].append([i1, i2, i3, i4, i5, i6, i7, i8])
            if n >= 9 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                for i6 in range(i5 + 1, n) :
                                    for i7 in range(i6 + 1, n) :
                                        for i8 in range(i7 + 1, n) :
                                            for i9 in range(i8 + 1, n) :
                                                combi[-1].append([i1, i2, i3, i4, i5, i6, i7, i8, i9])
            if n >= 10 : combi.append([])
            for i1 in range(n) :
                for i2 in range(i1 + 1, n) :
                    for i3 in range(i2 + 1, n) :
                        for i4 in range(i3 + 1, n) :
                            for i5 in range(i4 + 1, n) :
                                for i6 in range(i5 + 1, n) :
                                    for i7 in range(i6 + 1, n) :
                                        for i8 in range(i7 + 1, n) :
                                            for i9 in range(i8 + 1, n) :
                                                for i10 in range(i9 + 1, n) :
                                                    combi[-1].append([i1, i2, i3, i4, i5, i6, i7, i8, i9, i10])
            for i in range(k) :
                for j in range(len(combi[i])) :
                    combis[-1].append(combi[i][j])
    return combis
    
def get_nomenclature(wb_name : str) -> list :
    wb = load_workbook(filename = wb_name)
    ws = wb['Nomenclature SolidWorks']
    last_r = len(ws['A'])
    last_c = len(ws[1])
    
    nomenclature = list()
    for r in range(1, last_r) : 
        if ws.cell(row = r + 1, column = 1).value != None \
        and ws.cell(row = r + 1, column = 3).value in ["Profil", "Joint"] :
            nomenclature.append([])
            for c in range(2, last_c) :
                val = ws.cell(row = r + 1, column = c + 1).value
                if type(val) == str:
                    val = val.replace('\n', '')
                nomenclature[-1].append(val)                    
    return nomenclature

def get_conditionnements(wb_name : str) -> list :
    wb = load_workbook(filename = wb_name)
    ws = wb['Conditionnements']
    last_r = len(ws['A'])
    last_c = len(ws[1])
    
    for r in range(last_r + 1) :
        if ws.cell(row = r + 1, column = 1).value == None :
            break
    last_r = r
    
    conditionnements = [["" for _ in range(last_c)] for _ in range(last_r - 1)]
    for r in range(1, last_r) : 
        for c in range(last_c) :
            if ws.cell(row = r + 1, column = c + 1).value == None :
                conditionnements[r - 1][c] = ""
            else :
                conditionnements[r - 1][c] = ws.cell(row = r + 1, column = c + 1).value
    return conditionnements
def get_accessoires(wb_name : str) -> list :
    wb = load_workbook(filename = wb_name)
    ws = wb['Accessoires']
    last_r = ws.max_row + 1
    last_c = ws.max_column + 1
    
    accessoires = list()
    for c in range(1, last_c) : 
        if ws.cell(1, c).value != None :
            accessoires.append([])
            accessoires[-1].append(ws.cell(1, c).value)
            accessoires[-1].append([])
            for r in range(2, last_r) :
                if ws.cell(r, c).value != None :
                    accessoires[-1][-1].append([ws.cell(r, c).value, ws.cell(r, c + 1).value])
    return accessoires

def get_besoins(nomenclature : list) -> list :
    refs = set()
    besoins = list()
    for n1 in range(len(nomenclature)) :
        nom1 = nomenclature[n1]
        if nom1[0] == "Profil" or nom1[0] == "Joint":
            ref1 = nom1[1]
            if ref1 not in refs :
                refs.add(ref1)
                besoins.append([])
                besoins[-1].append(nom1)
                for n2 in range(len(nomenclature)) :                
                    nom2 = nomenclature[n2]
                    ref2 = nom2[1]
                    if n1 != n2 and ref1 == ref2 :
                        besoins[-1].append(nom2)
    return besoins

def shape_besoins(besoins : list) -> list :
    #Split quantités > 1
    besoins_shaped = list()
    for besoin in besoins :
        besoins_shaped.append([])
        for bes in besoin :
            bes_clone = bes.copy()            
            bes_clone[3] = 1
            for b in range(bes[3]) :
                besoins_shaped[-1].append(bes_clone)
    #Split longueurs > cond_max
    for b1 in range(len(besoins)) :
        # cond_max = conds[b1][-1][3] * 1000 
        cond_max = 6500
        for b2 in range(len(besoins_shaped[b1])) :
            longueur = besoins_shaped[b1][b2][4]
            nb_split = int(longueur / cond_max) + 1            
            besoins_shaped[b1][b2][4] = round(longueur / nb_split, 1)
            for s in range(nb_split - 1) :
                besoins_shaped[b1].append(besoins_shaped[b1][b2])
    #Tri
    besoins_sorted = list()
    for besoin_shaped in besoins_shaped :
        besoins_sorted.append(sorted(besoin_shaped, key=lambda x: x[4]))
    return besoins_sorted

def get_conds(besoins : list, conditionnements : list) -> list :
    conds = list()
    for besoin in besoins :
        ref = str(besoin[0][1])[:4]
        conds.append([])
        for cond in conditionnements :
            if str(cond[2]).find(str(ref)) > -1 :
                conds[-1].append(cond)
    conds_sorted = list()
    for cond in conds :
        conds_sorted.append(sorted(cond, key=lambda x: x[3]))
    return conds_sorted

def get_besoins_max(besoins_shaped : list, conds : list) -> list :
    besoins_max = list()
    
    for b1 in range(len(besoins_shaped)) :
        ref = besoins_shaped[b1][0][1]
        cond_max = conds[b1][-1][4] * 1000
        # print("\n cond max : \n", ref, cond_max)
        besoin_somme = 0
        count = 0
        for b2 in range(len(besoins_shaped[b1])) :
            if besoin_somme + besoins_shaped[b1][b2][4] < cond_max :
                besoin_somme += besoins_shaped[b1][b2][4]
                count = count + 1
            else :
                break
        besoins_max.append([ref, max(1, count)])
    return besoins_max

def get_combis_opti(combis : list, besoins_shaped : list, besoins_shaped_check : list, conds : list) -> list :
    combis_opti = list()
    for c in range(len(conds)) :
    # for c in range(2, 3) : 
        # print(conds[c])
        #pprint(besoins_shaped[c])
        #pprint(combis[c]) ----> VIDE !!!
        
        combis_opti.append([])
        while True in besoins_shaped_check[c] :
            chute = 100000
            # print(conds[c])
            for cond in conds[c] :
                # print("\n", cond)
                if cond[1] == "Joint" :
                    for combi in combis[c] :
                        best = [cond[2]] + combi
                else :
                    longueur = cond[4] * 1000
                    for combi in combis[c] :                    
                        combi_longueurs = [besoins_shaped[c][k][4] for k in combi]
                        if longueur - sum(combi_longueurs) >= 0 and longueur - sum(combi_longueurs) < chute :
                            chute = longueur - sum(combi_longueurs)
                            best = [cond[2]] + combi
            combis_opti[-1].append(best)
            # print(best)
            # print("c : ")
            for index_besoin in sorted(combis_opti[-1][-1][1:], reverse = True) :
                besoins_shaped_check[c][index_besoin] = False
            
            combis_clone = list()
            for d in range(len(combis)) :
                if d == c :
                    combis_clone.append([]) 
                else :
                    combis_clone.append(combis[d])
            for combi in combis[c] :
                ok = True
                for i in combis_opti[-1][-1][1:] :
                    if i in combi :
                        ok = False 
                        break
                if ok == True :
                    combis_clone[c].append(combi)
            combis = combis_clone.copy()
    return combis_opti

def get_new_wb_name(wb_name, path) :
    wb = load_workbook(filename = wb_name)
    ws = wb['SuperBar']
    ref = str(ws['B3'].value)
    
    now = datetime.datetime.now()
    return path \
        + '2_resultats\\' \
        + ref.replace(" ", "_") + "_" \
        + str(now.year).zfill(4) + "_" + str(now.month).zfill(2) + "_" + str(now.day).zfill(2) + "_" \
        + str(now.hour).zfill(2) + "_" + str(now.minute).zfill(2) + "_" + str(now.second).zfill(2) \
        + ".xlsx"

# def get_new_wb_name(wb_name) :
#     wb = load_workbook(filename = wb_name)
#     ws = wb['SuperBar']
#     ref = str(ws['B3'].value)
    
#     now = datetime.datetime.now()
#     return "BC_" \
#         + ref.replace(" ", "_") + "_" \
#         + str(now.year).zfill(4) + "_" + str(now.month).zfill(2) + "_" + str(now.day).zfill(2) + "_" \
#         + str(now.hour).zfill(2) + "_" + str(now.minute).zfill(2) + "_" + str(now.second).zfill(2) \
#         + ".xlsx"
        
def create_wb(wb_name) :
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Debit atelier"
    wb.create_sheet(title = "Bon de commande")
    wb.save(wb_name)
    return wb

def copy_sheet_bc(wb, new_wb, new_wb_name, new_ws_name, path) :
    wb_sh, new_wb_sh = wb["Bon de commande"], new_wb[new_ws_name]
    mr = wb_sh.max_row
    mc = wb_sh.max_column
    new_wb_sh.set_printer_settings(wb_sh.paper_size, wb_sh.orientation)
    new_wb_sh.page_margins = wb_sh.page_margins
    new_wb_sh.oddHeader = wb_sh.oddHeader
    new_wb_sh.oddFooter= wb_sh.oddFooter
    # img = drawing.image.Image(path + 'Logo.png') 
    # img.height = 100
    # img.width = 100
    # new_wb_sh.add_image(img, 'E3')
    new_wb_sh['E3'].alignment = styles.Alignment(horizontal='left')
    new_wb.save(filename = new_wb_name)
    
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            new_wb_sh.column_dimensions[utils.get_column_letter(j)].width \
                = wb_sh.column_dimensions[utils.get_column_letter(j)].width
            cell = wb_sh.cell(row = i, column = j)
            new_cell = new_wb_sh.cell(row = i, column = j)
            new_cell.value = cell.value
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
    new_wb.save(filename = new_wb_name)
    
def get_bcs(combis, conditionnements) :
    bcs = list()
    for combi in combis :
        ref = combi[0][0]
        if combi[0][1] == 'A' :
            qte = ' '
        else :
            qte = len(combi)
        categorie, designation, cdt = '', '', ''
        for cond in conditionnements :
            if cond[2] == ref :
                famille, categorie, designation, cdt = cond[0], cond[1], cond[3], str(cond[4]) + ' ' + str(cond[5])
                break
        bcs.append([famille, categorie, ref, designation, cdt, qte])
    return sorted(sorted(bcs, key=lambda x: x[1], reverse = True), key=lambda x: x[0])

def write_table(new_wb, new_wb_name, tab, ws_name) :
    ws = new_wb[ws_name]
    ws.cell(1, 1).value = ws_name.upper()
    nb, nb_global = 1, 0
    fam = ''
    fusions = list()
    for i in range(len(tab)) :
        if tab[i][1] != '' :
            for j in range(len(tab[i])) :
                
                ws.cell(i + 14, j + 1).value = tab[i][j]
                ws.cell(i + 14, j + 1).alignment = styles.Alignment(horizontal='center')
                ws.cell(i + 14, j + 1).border = styles.borders.Border(left=Side(style='thin'), \
                                                                     right=Side(style='thin'), \
                                                                     top=Side(style='thin'), \
                                                                     bottom=Side(style='thin'))
            
            if ws_name == "Debit atelier" : 
                f = '"STOCK'
                if tab[i][1] in ['Profil', 'Joint'] :
                    for k in range(1, int(tab[i][5]) + 1):
                        f += ',' + str(k)
                f += '"'
                # print(f)
                dv = DataValidation(type="list", formula1=f)
                dv.ranges.add('F' + str(i + 14))
                ws.add_data_validation(dv)
                
                if i == len(tab) - 1 :
                    end_index = i + 14
                else :
                    for k in range(i + 1, len(tab)) :
                        if tab[k][1] != '' or k == len(tab) - 1 :
                            end_index = k + 14 - 1
                            break
                ws.conditional_formatting.add('B' + str(i + 14) + ':F' + str(end_index), \
                                      FormulaRule(formula=['$F$' + str(i + 14) + '=\"STOCK\"'], \
                                      fill=styles.PatternFill(bgColor="F2DCDB", fill_type = "solid")))
            else :
                if i == len(tab) - 1 :
                    end_index = i + 14
                else :
                    for k in range(i + 1, len(tab)) :
                        if tab[k][1] != '' or k == len(tab) - 1 :
                            end_index = k + 14 - 1
                            break
                ws.conditional_formatting.add('B' + str(i + 14) + ':F' + str(end_index), \
                                      FormulaRule(formula=['$F$' + str(i + 14) + '=0'], \
                                      fill=styles.PatternFill(bgColor="F2DCDB", fill_type = "solid")))
                if tab[i][1] != 'Accessoire' :
                    ws[str(utils.get_column_letter(j+1)) + str(i+14)] = \
                            "=SUMIF(\'Debit atelier\'!$C:$C," \
                                    + "\'Bon de commande\'!$C" + str(i + 14) + "," \
                                    + "\'Debit atelier\'!$F:$F)"
        else :
            for j in range(len(tab[i])) :
                ws.cell(i + 14, j + 1).value = tab[i][j]
                ws.cell(i + 14, j + 1).font = styles.Font(color = "4F81BD", italic = True)
                if j in [2, 4] :
                    ws.cell(i + 14, j + 1).alignment = styles.Alignment(horizontal='right')
                elif j in [3, 5] :
                    ws.cell(i + 14, j + 1).alignment = styles.Alignment(horizontal='left')
                else :
                    ws.cell(i + 14, j + 1).alignment = styles.Alignment(horizontal='center')
        
        if tab[i][0] != fam :
            if fam != '' :
                fusions.append([fam, 14 + nb_global, 14 + nb_global - 1 + nb - 1])
            nb_global += nb - 1
            nb = 1
            fam = tab[i][0]
        nb += 1
    
            
    fusions.append([fam, 14 + nb_global, 14 + nb_global - 1 + nb - 1])
    # pprint(fusions)
    
    for fusion in fusions :
        ws.merge_cells(start_row=fusion[1], start_column=1, end_row=fusion[2], end_column=1)
        ws.cell(fusion[1], 1).alignment = styles.Alignment(textRotation=90, \
                                                           horizontal = 'center', \
                                                           vertical = 'center', \
                                                           wrap_text=True)
        ws.cell(fusion[1], 1).font = styles.Font(size = "22", bold = True)
        if fusion[0] == "Chevrons" :
            ws.cell(fusion[1], 1).fill = styles.PatternFill(start_color ="FFF2CC", fill_type = "solid")
        elif fusion[0] == "Chéneau" :
            ws.cell(fusion[1], 1).fill = styles.PatternFill(start_color="E2EFDA", fill_type = "solid")
        elif fusion[0] == "Faitage" :
            ws.cell(fusion[1], 1).fill = styles.PatternFill(start_color="DDEBF7", fill_type = "solid")
        elif fusion[0] == "Poteaux" :
            ws.cell(fusion[1], 1).fill = styles.PatternFill(start_color="FCE4D6", fill_type = "solid")
    new_wb.save(filename = new_wb_name)
    
def fill_accessoires(wb_name, accessoires, combis_opti) :
    wb = load_workbook(filename = wb_name)
    type_veranda = wb['SuperBar']['B4'].value
    for accessoire in accessoires :
        if accessoire[0] == type_veranda :
            for access in accessoire[1] :
                combis_opti.append([[access[0], "A"]])
            break
    return combis_opti

def get_ats(bcs, combis_opti, besoin_shaped, conds) :
    ats = list()
    for bc in bcs :  
        fam, ref = bc[0], bc[2]  
        if bc[1] == 'Accessoire' :
            ats.append(bc)
        else :
            for c1 in range(len(combis_opti)) :
                combi = combis_opti[c1]
                if combi[0][0] == ref :
                    for c2 in range(len(combi)) :
                        comb = combi[c2]
                        ats.append(bc)
                        ats[-1][5] = 1
                        longueur = 0  
                        nb = 0
                        count = 0
                        for i in range(1, len(comb)) :
                            nb += 1 
                            longueur += besoin_shaped[c1][comb[i]][4]
                            if (i == len(comb) - 1 and len(comb) == 2) \
                            or (i == len(comb) - 1 \
                                and besoin_shaped[c1][comb[i]][4] == besoin_shaped[c1][comb[i - 1]][4] \
                                and besoin_shaped[c1][comb[i]][5] == besoin_shaped[c1][comb[i - 1]][5] \
                                and besoin_shaped[c1][comb[i]][6] == besoin_shaped[c1][comb[i - 1]][6]) : 
                                count += 1
                                ats.append(['' for _ in range(6)])
                                ats[-1][0] = fam
                                ats[-1][2] = 'Coupe x' + str(nb) + " : "
                                ats[-1][3] = str(besoin_shaped[c1][comb[i]][4]) \
                                            + ' | ' + str(besoin_shaped[c1][comb[i]][5]) \
                                            + ' | ' + str(besoin_shaped[c1][comb[i]][6]) 
                                nb = 1
                            elif i > 1 \
                            and (besoin_shaped[c1][comb[i]][4] != besoin_shaped[c1][comb[i - 1]][4] \
                            or besoin_shaped[c1][comb[i]][5] != besoin_shaped[c1][comb[i - 1]][5] \
                            or besoin_shaped[c1][comb[i]][6] != besoin_shaped[c1][comb[i - 1]][6]) :
                                count += 1
                                ats.append(['' for _ in range(6)])
                                ats[-1][0] = fam
                                ats[-1][2] = 'Coupe x' + str(nb - 1) + " : "
                                ats[-1][3] = str(besoin_shaped[c1][comb[i - 1]][4]) \
                                            + ' | ' + str(besoin_shaped[c1][comb[i - 1]][5]) \
                                            + ' | ' + str(besoin_shaped[c1][comb[i - 1]][6]) 
                                
                                if i == len(comb) - 1 :
                                    count += 1
                                    ats.append(['' for _ in range(6)])
                                    ats[-1][0] = fam
                                    ats[-1][2] = 'Coupe x' + str(nb - 1) + " : "
                                    ats[-1][3] = str(besoin_shaped[c1][comb[i]][4]) \
                                                + ' | ' + str(besoin_shaped[c1][comb[i]][5]) \
                                                + ' | ' + str(besoin_shaped[c1][comb[i]][6]) 
                                nb = 1
                        ats[-count][4] = 'Total : '
                        ats[-count][5] = longueur
                        if ats[-count-1][1] == 'Joint' :
                            if ats[-count-1][4].find('25') > -1 :
                                ats[-count-1][5] = int(longueur / 25000) + 1
                            elif ats[-count-1][4].find('50') > -1 :
                                ats[-count-1][5] = int(longueur / 50000) + 1
    return ats
        
path = 'Z:\\Commun\\2. PARTIE BE\\superbar\\'
wb_name = path + '0_sources\\SuperBar_tmp.xlsm'
wb = load_workbook(filename = wb_name, data_only = True)
ws_name = 'SuperBar'

print("\n\n*******************SUPERBAR*******************")
print("\nRécupération de la nomenclature...")
nomenclature = get_nomenclature(wb_name)
print("OK")
# pprint(nomenclature)
print("\nRécupération des conditionnements...")
conditionnements = get_conditionnements(wb_name)
print("OK")
# pprint(conditionnements)
print("\nRécupération des accessoires...")
accessoires = get_accessoires(wb_name)
# pprint(accessoires)
print("OK")
print("\nCalcul des besoins...")
besoins = get_besoins(nomenclature)
# print(len(besoins))
# pprint(besoins)
print("OK")
print("\nMise en forme des besoins...")
besoins_shaped = shape_besoins(besoins)
# print(len(besoins_shaped))
# pprint(besoins_shaped)
print("OK")
print("\nMise en forme des conditionnements...")
conds = get_conds(besoins, conditionnements)
# print(len(conds))
# pprint(conds)
print("OK")
print("\nCalcul des besoins max...")
besoins_max = get_besoins_max(besoins_shaped, conds)
# print(len(besoins_max))
# pprint(besoins_max)
print("OK")
print("\nCréation du tableau de checks besoins...")
besoins_shaped_check = [[True for _ in range(len(besoins_shaped[b]))] for b in range(len(besoins_shaped))]
# print(len(besoins_shaped_check))
# pprint(besoins_shaped_check)
print("OK")
print("\nCalcul des combinaisons...")
combis = get_combis(besoins_shaped, besoins_max)
# pprint(combis)
print("OK")
print("\nCalcul des combinaisons optimales...")
combis_opti = get_combis_opti(combis, besoins_shaped, besoins_shaped_check, conds)
# pprint(combis_opti)
print("OK")
print("\nAjout des accessoires...")
combis_opti = fill_accessoires(wb_name, accessoires, combis_opti)
# pprint(combis_opti)
print("OK")
print("\nCalcul du nom du nouveau classeur...")
new_wb_name = get_new_wb_name(wb_name, path)
# print(new_wb_name)
print("OK")
print("\nCréation du nouveau classeur...")
new_wb = create_wb(new_wb_name)
print("OK")
print("\nCopie / création de la feuille bon de commande...")
copy_sheet_bc(wb, new_wb, new_wb_name, "Bon de commande", path)
print("OK")
print("\nCopie / création de la feuille débit atelier...")
copy_sheet_bc(wb, new_wb, new_wb_name, "Debit atelier", path)
print("OK")
print("\nMise en forme des lignes bon de commande...")
bcs = get_bcs(combis_opti, conditionnements)
# pprint(bcs)
print("OK")
print("\nEcriture des lignes bon de commande...")
write_table(new_wb, new_wb_name, bcs, "Bon de commande")
print("OK")
print("\nMise en forme des lignes débit atelier...")
ats = get_ats(bcs, combis_opti, besoins_shaped, conds)
# pprint(ats)
print("OK")
print("\nEcriture des lignes débit atelier...")
write_table(new_wb, new_wb_name, ats, "Debit atelier")
print("OK")


# wb = Workbook()
# ws = wb.create_sheet("new")
# ws['A1'] = "ok"
# wb.save("new wb.xlsx")
